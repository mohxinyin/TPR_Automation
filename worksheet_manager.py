import constants as c
import datetime
from copy import copy
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from file_handler import load_excel_workbook

def prepare_working_sheet(wb, header_wb, new_sheet_name, header_sheet_name, cols_to_delete, source_sheet_name = "Sheet1" ):
    """
    Takes an openpyxl Workbook object, copies a sheet, hides the original,
    renames the copy, and adds a styled header from another workbook.

    Parameters:
    - wb (Workbook): Already loaded main workbook.
    - header_wb (Workbook): Already loaded header workbook.
    - source_sheet_name (str): Name of the sheet to copy.
    - new_sheet_name (str): Desired name for the copied sheet.

    Returns:
    - Workbook: Modified workbook object.
    """
    # Copy and rename the sheet
    source_ws = wb[source_sheet_name]
    working_ws = wb.copy_worksheet(source_ws)
    working_ws.title = new_sheet_name

    # Hide the original sheet 
    source_ws.sheet_state = 'hidden'

    # Get the header row from the header workbook
    header_ws = header_wb[header_sheet_name]
    working_ws.insert_rows(2)

    # Copy the header style from TPR Header 
    copy_header_styles(header_ws,wb,header_row = 2)
    print(f"'{new_sheet_name}' sheet has been created with header.")

    remove_unwanted_columns(working_ws,cols_to_delete)

    # Adjust width of column to max length
    adjust_column_width(wb)
    return wb

def remove_unwanted_columns(ws, cols_to_delete):
    """
    Removes unwanted columns from the worksheet based on COLUMNS_TO_DELETE.
    Deletion is done from right to left to avoid shifting.
    """
    # Convert letters to indices, remove duplicates, and sort descending
    cols_to_remove = sorted(
        {column_index_from_string(col) for col in cols_to_delete},
        reverse=True
    )

    for col_idx in cols_to_remove:
        ws.delete_cols(col_idx)

    print("Selected columns removed.")

    if ws.title == 'Working' or ws.title == "TPR Working":
        ws.delete_rows(1) # Remove first row from working sheets 
        print("Top row removed from working sheet.")

def adjust_column_width(wb):
    # Make sure all values are visible --> adjust width of column to max length and freeze top row
    for sheet in wb.worksheets:

        # Freeze top pane 
        sheet.freeze_panes = 'A2'

        for col in sheet.columns:
            max_length = 0  # Track the maximum length of the value in the column
            column = col[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)

            for cell in col:
                try:
                    # Only consider the value of the cell (ignoring the header)
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Adjust the column width to the max length found
            adjusted_width = (max_length + 2)  # Add 2 to give some padding space
            sheet.column_dimensions[column].width = adjusted_width

def copy_header_styles(styled_ws, wb,header_row):
    """
    Copies the header styles from the styled_ws (reference) to all sheets in the workbook (wb).
    
    Parameters:
    - styled_ws: The worksheet containing the reference header styles
    - wb: The workbook containing the sheets to apply the header style to
    """
    for sheet_name in wb.sheetnames:
        green_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        # Skip the sheet named 'Sheet1' and 'Inventory by WH'
        if sheet_name in ['Sheet1','Inventory by WH','Summary']:
            continue

        sheet = wb[sheet_name]  # Access each sheet by name

        # Only apply the style to the first row (header)
        for col_index, value in enumerate(styled_ws[1], start=1):
            target_cell = sheet.cell(row=header_row, column=col_index)
            target_cell.value = value.value
            target_cell.fill = green_fill           
            target_cell.font = Font(bold=True)

            # Copy styles from reference header
            if value.has_style:
                target_cell.border = copy(value.border)
                target_cell.alignment = copy(value.alignment)
                target_cell.number_format = value.number_format
                target_cell.protection = copy(value.protection)

def create_new_columns(ws, new_headers, after_col_letter = None):
    for col in range(ws.max_column, 0, -1):
        if ws.cell(row=1, column=col).value not in (None, ''):
            last_col_idx = col
            break
    bold_font = Font(bold = True)

    for i, header in enumerate(new_headers):
        if ws.title == "Inventory by WH":
            after_col_idx = column_index_from_string(after_col_letter)
            insert_at = after_col_idx + i + 1
        else:
            insert_at = last_col_idx + i + 1

        # Skip an additional column after 'Suggestion' in the Summary sheet 
        if header == 'Suggestion':  # When 'Suggestion' column is reached, skip the next column
            last_col_idx += 1

        ws.insert_cols(insert_at)
        cell = ws.cell(row=1, column=insert_at)
        cell.value = header

        if ws.title == 'Schedule':
            cell.font = bold_font  # Apply bold font

def add_year_month_columns(ws, due_date_col=c.due_date_idx):
    """
    Adds Year and Month columns based on the Due Date at the end of the existing columns.
    Populates Year and Month values.
    """
    # Find last row with data in the due date column
    last_row = ws.Cells(ws.Rows.Count, due_date_col).End(-4162).Row  # xlUp = -4162

    # Find last used column in the first row
    last_col = ws.Cells(1, ws.Columns.Count).End(-4159).Column  # xlToLeft = -4159

    # Decide where to insert the Year and Month
    year_col = last_col + 1
    month_col = last_col + 2

    # Set headers
    ws.Cells(1, year_col).Value = "Year"
    ws.Cells(1, month_col).Value = "Month"

    # Fill values
    for i in range(2, last_row + 1):
        due_date = ws.Cells(i, due_date_col).Value
        if isinstance(due_date, datetime.datetime):
            ws.Cells(i, year_col).Value = due_date.year
            ws.Cells(i, month_col).Value = due_date.strftime("%b")
        else:
            ws.Cells(i, year_col).Value = None
            ws.Cells(i, month_col).Value = None

    print(f"Year and Month columns added to the Schedule sheet at columns {year_col} and {month_col}.")


def import_inventory_sheet(source_file_path, target_wb, source_sheet_name = 'Results', new_sheet_name='Inventory by WH', before_sheet_name = 'MRP'):
    # Load the source workbook and sheet
    src_wb = load_excel_workbook(source_file_path)
    src_ws = src_wb[source_sheet_name]

    # Create a new sheet in the target workbook
    target_ws = target_wb.create_sheet(title=new_sheet_name)

    # Copy content from source to target
    for row in src_ws.iter_rows():
        for cell in row:
            target_ws.cell(row=cell.row, column=cell.column).value = cell.value

    # Reorder sheets to put new sheet before 'MRP'
    sheets = target_wb._sheets
    mrp_index = sheets.index(target_wb[before_sheet_name])
    new_sheet_index = sheets.index(target_ws)

    # Remove and insert in the correct position
    sheets.insert(mrp_index, sheets.pop(new_sheet_index))

    print(f"{source_sheet_name} copied to target workbook as {new_sheet_name} and {new_sheet_name} inserted before {before_sheet_name}.")

def create_summary_sheet(wb):
    green_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    bold_font = Font(bold=True)

    # Access 'OHS' sheet
    ohs_ws = wb['OHS']

    # Create new sheet and insert before 'OHS'
    idx = wb.sheetnames.index('OHS')
    summary_ws = wb.create_sheet(title="Summary", index=idx)

    # Copy columns A to G from 'OHS' to 'Summary'
    for row_idx in range(1, ohs_ws.max_row + 1):
        new_col_idx = 1  # Start placing in column A in 'Summary'
        for col_idx in c.COLUMNS_TO_COPY_SUMMARY:
            cell = ohs_ws.cell(row=row_idx, column=col_idx)
            new_cell = summary_ws.cell(row=row_idx, column=new_col_idx, value=cell.value)
            if cell.has_style:
                new_cell._style = cell._style
            # Apply green fill to the header row (row 1)
            if row_idx == 1:  # Apply only to header row
                new_cell.fill = green_fill
                new_cell.font = bold_font 
            new_col_idx += 1

    print("Summary sheet created and columns A to G and P copied from OHS.")
    # Rename column H to 'On-hand Stock'
    summary_ws['H1'].value = 'On-hand Stock'
    return wb

def format_due_date(wb,due_date_idx):
    for ws in wb.worksheets:
        if ws.title == 'Summary':
            continue
        # Apply 'DD/MM/YYYY' format to all rows in column J('Due Date')
        for row in ws.iter_rows(min_row=2, min_col=due_date_idx, max_col=due_date_idx):
            for cell in row:
                cell.number_format = 'DD/MM/YYYY'




            





