import pandas as pd
import constants as c 

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import column_index_from_string

def filter_and_create_sheet(wb, filters, output_sheet_name, source_sheet_name):
    """
    Filters data in a worksheet and creates a new sheet with filtered results.

    Parameters:
    - wb: openpyxl Workbook object
    - source_sheet_name: name of the sheet to filter
    - filters: list of functions that return True/False given a DataFrame
    - output_sheet_name: name of the new worksheet to create with filtered results
    """
    ws = wb[source_sheet_name]
    
    # Convert worksheet to DataFrame
    data = ws.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)

    # Apply filters
    for condition in filters:
        df = df[condition(df)]

    # Remove sheet if it already exists
    if output_sheet_name in wb.sheetnames:
        del wb[output_sheet_name]

    # Create new sheet and write the filtered data
    new_ws = wb.create_sheet(title=output_sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        new_ws.append(r)

    print(f"Filtered data written to '{output_sheet_name}' sheet.")
    return df

def fill_column_based_on_filter(ws, target_col_letter, condition_func, source_col_letter = 'I', label='Y'):
    """
    Fill a column with 'Y' based on a filter applied to the Source column.

    Args:
        ws: The worksheet object.
        source_col_letter (str): Column letter for 'Source' (e.g., 'I').
        target_col_letter (str): Column letter to write the 'Y' (e.g., 'P', 'Q', etc.).
        condition_func (function): A function that takes a cell value and returns True if the row should be updated.
        label (str): The value to insert in the target column (default 'Y').
    """
    source_col_idx = column_index_from_string(source_col_letter)
    target_col_idx = column_index_from_string(target_col_letter)

    updated_rows = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        source_cell_val = row[source_col_idx - 1].value
        target_cell = row[target_col_idx - 1]

        if condition_func(str(source_cell_val).strip() if source_cell_val else ''):
                target_cell.value = label
                updated_rows += 1

    print(f"{updated_rows} cells in column {target_col_letter} filled with '{label}'.")


     