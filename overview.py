import constants as c
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from worksheet_manager import create_new_columns
from api_handler import update_excel_with_standard_cost

def create_overview_sheet(wb):
    green_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    bold_font = Font(bold=True)
    
    # COPY COLUMNS A-F FROM OHS SHEETS
    # Access 'OHS' sheet
    ohs_ws = wb['OHS']

    # Create new sheet and insert before 'OHS'
    idx = wb.sheetnames.index('OHS')
    overview_ws = wb.create_sheet(title="Overview", index=idx)

    # Copy columns A to F,K from 'OHS' to 'Overview'
    for row_idx in range(1, ohs_ws.max_row + 1):
        new_col_idx = 1  # Start placing in column A in 'Overview'
        for col_idx in c.COLUMNS_TO_COPY_OVERVIEW:
            cell = ohs_ws.cell(row=row_idx, column=col_idx)
            new_cell = overview_ws.cell(row=row_idx, column=new_col_idx, value=cell.value)
            if cell.has_style:
                new_cell._style = cell._style
            new_col_idx += 1

    print("Overview sheet created and columns A to F and K copied from OHS.")

    # Rename column G to 'Inventory'
    overview_ws['G1'].value = 'Inventory'
    # overview_ws = wb['Overview']
    
    # ADD HEADERS AFTER COLUMN G
    create_new_columns(overview_ws,c.COLUMNS_TO_ADD_OVERVIEW)

    # COPY COLUMN 'P'(BalanceQty2) from ohs to COLUMN 'Q' (Proj) of overview
    for row_idx in range(1, ohs_ws.max_row + 1):
        source_cell = ohs_ws.cell(row=row_idx, column=column_index_from_string('P'))  # Column p
        target_cell = overview_ws.cell(row=row_idx, column=column_index_from_string('Q'))  # Column Q = 17
        target_cell.value = source_cell.value

    overview_ws['Q1'].value = 'Proj'

    update_excel_with_standard_cost(wb,c.dest_file)

    # Find Inventory column index dynamically
    inventory_col_idx = None
    for cell in overview_ws[1]:  # Header row
        if cell.value and str(cell.value).strip().lower() == 'inventory':
            inventory_col_idx = cell.column
            break

    if inventory_col_idx:
        for row_idx in range(2, ohs_ws.max_row + 1):  # Use source data rows
            cell = overview_ws.cell(row=row_idx, column=inventory_col_idx)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = 0
    else:
        print("⚠ Inventory column not found!")

    # Apply green fill and bold font to header row (Row 1)
    for cell in overview_ws[1]:
        cell.fill = green_fill
        cell.font = bold_font

    return wb