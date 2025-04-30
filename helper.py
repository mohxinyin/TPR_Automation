from filtering import filter_and_create_sheet
from filtering import fill_column_based_on_filter
from openpyxl.utils import column_index_from_string
import constants as c 

def create_filtered_sheets(wb,sheet_config, source_sheet_name):
    for config in sheet_config():
        filter_and_create_sheet(
            wb=wb,
            filters=config["filters"],
            output_sheet_name=config["name"],
            source_sheet_name = source_sheet_name 
        )

def tpr_sheet_config():
    yield {
        "name": "MRP",
        "filters": [
            lambda df: df['Source'].str.contains('MRP', na=False, case=False),
            lambda df: df['Due Date'].dt.year == 2025,
            lambda df: df['Receipts'].notna() & (df['Receipts'].str.strip() != '')
        ]
    }
    yield {
        "name": "Schedule",
        "filters": [
            lambda df: df['Source'].str.contains('Job', na=False, case=False),
            lambda df: df['Type'] == 'M',
            lambda df: df['Receipts'].notna() & (df['Receipts'].str.strip() != '')
        ]
    }
    yield {
        "name": "TPR Inventory",
        "filters": [
            lambda df: df['Source'].str.contains('On-Hand Quantity', na=False, case=False)
        ]
    }
    print("Filtered sheets created")

def summary_sheet_config():
    yield {
        "name": "OHS",
        "filters": [
            lambda df: df['Source'].str.contains('On-hand', na=False, case=False),
        ]
    }
    yield {
        "name": "MO",
        "filters": [
            lambda df: df['Source'].str.startswith('Job', na=False,)
        ]
    }
    yield {
        "name": "SO",
        "filters": [
            lambda df: df['Source'].str.contains('SO:', na=False, case=False)
        ]
    }
    yield {
        "name": "PO",
        "filters": [
            lambda df: df['Source'].str.contains('PO:', na=False, case=False),
        ]
    }
    yield {
        "name": "Forecast",
        "filters": [
            lambda df: df['Source'].str.contains('Forecast', na=False, case=False),
        ]
    }
    yield {
        "name": "Suggestion",
        "filters": [
            lambda df: df['Source'].str.contains('Suggestion', na=False, case=False),
        ]
    }
    print("Filtered sheets created")

def fill_schedule_values(ws):
    # MRP
    fill_column_based_on_filter(ws,'Q', lambda val: 'MRP' in val.upper()) # Start from Q because of the 2 extra columns added ("year" and "month" columns)
    # MO
    fill_column_based_on_filter(ws,'R', lambda val: val.startswith('Job') and 'MRP' not in val.upper())
    # Expedite
    fill_column_based_on_filter(ws,'S', lambda val: 'expedite' in val.lower())
    # Postpone
    fill_column_based_on_filter(ws,'T', lambda val: 'postpone' in val.lower())

def pivot_table_generator():
    yield{
        'sheet_name' : 'MRP',
        'table_range' : c.MRP_Table_Range,
        'pivot_table_location' :'O1',
        'row_field' : ['Class'],
        'data_field' : [('PartNum','count')]
    }
    yield{
        'sheet_name' : 'Schedule',
        'table_range' : c.Schedule_Table_Range,
        'pivot_table_location' : 'X1',
        'row_field' : ['Year','Month','Due Date'],
        'data_field' : [('MRP','count'),('MO','count'),('EXPEDITE','count'),('POSTPONE','count')],
        'filter_field' : ['Class']
    }
    yield{
        'sheet_name' : 'Inventory by WH',
        'table_range' : c.Inventory_Pivot_Range,
        'pivot_table_location' :'O1',
        'row_field' : ['Part Num'],
        'column_field': ['Area'],
        'data_field' : [('On Hand','sum')]
    }        

def convert_to_general(wb): 
    for sheet_name, columns in c.sheets_and_columns.items():
        ws = wb[sheet_name]
        for col_letter in columns:
            for row in ws.iter_rows(min_col=column_index_from_string(col_letter),
                                    max_col=column_index_from_string(col_letter),
                                    min_row=2,
                                    max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        cleaned = str(cell.value).strip()
                        
                        if cleaned == "":
                            cell.value = None  # Empty cell, treat as blank
                        else:
                            try:
                                if "." in cleaned:
                                    cell.value = float(cleaned)
                                else:
                                    cell.value = int(cleaned)
                            except ValueError:
                                cell.value = cleaned  # fallback if not a number






